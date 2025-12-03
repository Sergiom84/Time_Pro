from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file
from functools import wraps
from models.models import User, TimeRecord, EmployeeStatus, Center, Category, OvertimeEntry
from models.database import db
from werkzeug.security import generate_password_hash
from datetime import datetime, timedelta, date
from sqlalchemy.orm import joinedload
import os
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from routes.auth import admin_required
from utils.logging_utils import get_logger

export_bp = Blueprint("export", __name__, template_folder="../templates")

STATUS_GROUPS = {
    "Trabajado": ["Trabajado"],
    "Baja": ["Baja"],
    "Ausente": ["Ausente"],
    "Vacaciones": ["Vacaciones"]
}
CATEGORY_NONE_VALUES = {"sin categoria", "sin categoría", "-- sin categoría --"}



def calculate_pause_data_for_records(time_records):
    """Calcula el tiempo total de pausa y el detalle de pausas por registro de tiempo."""
    pause_seconds_by_record = {}
    pause_details = []

    for record in time_records:
        total_seconds = 0
        if getattr(record, "pauses", None):
            for pause in record.pauses:
                if not pause.pause_start or not pause.pause_end:
                    continue
                duration_seconds = int(
                    max((pause.pause_end - pause.pause_start).total_seconds(), 0)
                )
                total_seconds += duration_seconds
                pause_details.append({
                    "user": record.user,
                    "date": record.date,
                    "pause": pause,
                    "duration_seconds": duration_seconds,
                    "notes": pause.notes or "",
                })

        pause_seconds_by_record[record.id] = total_seconds

    return pause_seconds_by_record, pause_details


def format_duration_from_seconds(seconds):
    """Devuelve una cadena HH:MM a partir de segundos."""
    if seconds is None:
        return ""

    total_seconds = int(max(seconds, 0))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}"

def expand_status_filters(filters):
    """Expande filtros de estado generales a sus valores específicos."""
    expanded = []
    for status in filters:
        expanded.extend(STATUS_GROUPS.get(status, [status]))
    # Mantener orden y evitar duplicados
    seen = set()
    ordered_unique = []
    for status in expanded:
        if status not in seen:
            seen.add(status)
            ordered_unique.append(status)
    return ordered_unique


def resolve_category_filter(value):
    """Devuelve (category_id, filter_none_flag) para un nombre recibido en formularios."""
    if not value or value in ("", "all"):
        return None, False

    normalized = value.strip().lower()
    if normalized in CATEGORY_NONE_VALUES:
        return None, True

    client_id = session.get("client_id")
    if not client_id:
        return None, False

    category = Category.query.filter_by(client_id=client_id, name=value).first()
    return (category.id if category else None), False


def get_user_category_label(user, default="-"):
    """Obtiene el nombre legible de la categoría del usuario."""
    if not user:
        return default
    if user.category:
        return user.category.name
    return default


def add_overtime_sheet_to_workbook(wb, overtime_entries):
    """
    Añade una pestaña de Horas Extras al workbook.
    overtime_entries: lista de OvertimeEntry ya filtrados
    """
    if not overtime_entries:
        return

    ws = wb.create_sheet("Horas Extras")

    # Encabezados
    headers = [
        "Usuario",
        "Nombre",
        "Categoría",
        "Centro",
        "Semana (inicio)",
        "Semana (fin)",
        "Jornada Semanal",
        "Horas Trabajadas",
        "Horas Extra/Déficit",
        "Estado",
        "Decidido por",
        "Fecha Decisión",
        "Notas"
    ]

    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Datos
    row_num = 2
    for entry in overtime_entries:
        user = entry.user_rel
        decider = User.query.get(entry.decided_by) if entry.decided_by else None

        # Convertir segundos a horas
        worked_hours = entry.total_worked_seconds / 3600
        contract_hours = entry.contract_seconds / 3600
        overtime_hours = entry.overtime_seconds / 3600

        ws.cell(row=row_num, column=1).value = user.username if user else f"ID: {entry.user_id}"
        ws.cell(row=row_num, column=2).value = user.full_name if user else "-"
        ws.cell(row=row_num, column=3).value = get_user_category_label(user)
        ws.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
        ws.cell(row=row_num, column=5).value = entry.week_start.strftime("%d/%m/%Y")
        ws.cell(row=row_num, column=6).value = entry.week_end.strftime("%d/%m/%Y")
        ws.cell(row=row_num, column=7).value = f"{contract_hours:.2f}h"
        ws.cell(row=row_num, column=8).value = f"{worked_hours:.2f}h"
        ws.cell(row=row_num, column=9).value = f"{overtime_hours:+.2f}h"
        ws.cell(row=row_num, column=10).value = entry.status
        ws.cell(row=row_num, column=11).value = decider.username if decider else "-"
        ws.cell(row=row_num, column=12).value = entry.decided_at.strftime("%d/%m/%Y %H:%M") if entry.decided_at else "-"
        ws.cell(row=row_num, column=13).value = entry.decision_notes or "-"

        row_num += 1

    # Ajustar anchos de columna
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 17


# ========== FUNCIONES HELPER PARA EXPORTACIONES DIARIAS CON FILTROS ==========

def handle_daily_excel_export(req):
    """Maneja exportación Excel diaria con filtros de estado"""
    # Obtener fecha (hoy si no se especifica)
    start_date = req.form.get("start_date")
    end_date = req.form.get("end_date")

    if start_date:
        fecha = datetime.strptime(start_date, "%Y-%m-%d").date()
    else:
        fecha = date.today()

    # Obtener filtros de estado
    status_filters = list(dict.fromkeys(req.form.getlist('status')))
    if not status_filters:
        status_filters = ['Trabajado']

    # Obtener registros de TimeRecord (si "Trabajado" está seleccionado)
    time_records = []
    if 'Trabajado' in status_filters:
        # Eager loading: cargar User en la misma query para evitar N+1
        time_records = (
            TimeRecord.query
            .filter(TimeRecord.date == fecha)
            .options(joinedload(TimeRecord.user), joinedload(TimeRecord.pauses))
            .order_by(TimeRecord.check_in.desc())
            .all()
        )
        time_records = sorted(
            time_records,
            key=lambda r: r.check_in or datetime.combine(r.date, datetime.min.time()),
            reverse=True
        )

    # Obtener registros de EmployeeStatus (Baja, Ausente, Vacaciones)
    employee_statuses = []
    selected_statuses = expand_status_filters([s for s in status_filters if s != 'Trabajado'])
    if selected_statuses:
        employee_statuses = EmployeeStatus.query.filter(
            EmployeeStatus.status.in_(selected_statuses),
            EmployeeStatus.date == fecha
        ).order_by(EmployeeStatus.user_id).all()

    if not time_records and not employee_statuses:
        flash("No hay registros para ese día con los filtros seleccionados.", "warning")
        return redirect(url_for("export.export_excel"))


    # Calcular datos de pausas SOLO si el filtro "Pausas" está activo
    pause_seconds_by_record = {}
    pause_details = []
    if 'Pausas' in status_filters and time_records:
        pause_seconds_by_record, pause_details = calculate_pause_data_for_records(time_records)

    # Generar Excel con 3 pestañas
    wb = openpyxl.Workbook()

    # Pestaña 1: Registros de Fichaje
    if time_records:
        ws1 = wb.active
        ws1.title = "Registros de Fichaje"

        # Header condicional según filtro de Pausas
        if 'Pausas' in status_filters:
            header1 = [
                "Usuario",
                "Nombre completo",
                "Categoria",
                "Centro",
                "Fecha",
                "Entrada",
                "Salida",
                "Horas Totales",
                "Tiempo de Pausa",
                "Horas Efectivas",
                "Notas",
                "Notas Admin",
            ]
        else:
            header1 = [
                "Usuario",
                "Nombre completo",
                "Categoria",
                "Centro",
                "Fecha",
                "Entrada",
                "Salida",
                "Horas Trabajadas",
                "Notas",
                "Notas Admin",
            ]
        for col_num, header_text in enumerate(header1, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        row_num = 2
        for record in time_records:
            user = record.user
            total_seconds = None
            if record.check_in and record.check_out:
                total_seconds = int((record.check_out - record.check_in).total_seconds())

            # Pausas solo si filtro activo
            pause_seconds = pause_seconds_by_record.get(record.id, 0) if 'Pausas' in status_filters else 0
            effective_seconds = None
            if 'Pausas' in status_filters and total_seconds is not None:
                effective_seconds = max(total_seconds - pause_seconds, 0)

            # Escribir columnas comunes (1-7)
            ws1.cell(row=row_num, column=1).value = user.username if user else f"ID: {record.user_id}"
            ws1.cell(row=row_num, column=2).value = user.full_name if user else "-"
            ws1.cell(row=row_num, column=3).value = get_user_category_label(user)
            ws1.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
            ws1.cell(row=row_num, column=5).value = record.date.strftime("%d/%m/%Y")
            ws1.cell(row=row_num, column=6).value = record.check_in.strftime("%H:%M:%S") if record.check_in else "-"
            ws1.cell(row=row_num, column=7).value = record.check_out.strftime("%H:%M:%S") if record.check_out else "-"

            # Columnas condicionales según filtro
            if 'Pausas' in status_filters:
                ws1.cell(row=row_num, column=8).value = format_duration_from_seconds(total_seconds)
                ws1.cell(row=row_num, column=9).value = format_duration_from_seconds(pause_seconds)
                ws1.cell(row=row_num, column=10).value = format_duration_from_seconds(effective_seconds)
                ws1.cell(row=row_num, column=11).value = record.notes
                ws1.cell(row=row_num, column=12).value = record.admin_notes
            else:
                ws1.cell(row=row_num, column=8).value = format_duration_from_seconds(total_seconds)
                ws1.cell(row=row_num, column=9).value = record.notes
                ws1.cell(row=row_num, column=10).value = record.admin_notes

            row_num += 1

        for col_num, _ in enumerate(header1, 1):
            col_letter = get_column_letter(col_num)
            ws1.column_dimensions[col_letter].width = 17
    else:
        wb.remove(wb.active)

    # Pestaña 2: Bajas y Ausencias
    if employee_statuses:
        ws2 = wb.create_sheet("Bajas y Ausencias")
        header2 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Estado", "Notas", "Notas Admin"]
        for col_num, header_text in enumerate(header2, 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        row_num = 2
        for status_record in employee_statuses:
            user = User.query.get(status_record.user_id)
            ws2.cell(row=row_num, column=1).value = user.username if user else f"ID: {status_record.user_id}"
            ws2.cell(row=row_num, column=2).value = user.full_name if user else "-"
            ws2.cell(row=row_num, column=3).value = get_user_category_label(user)
            ws2.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
            ws2.cell(row=row_num, column=5).value = status_record.date.strftime("%d/%m/%Y")
            ws2.cell(row=row_num, column=6).value = status_record.status
            ws2.cell(row=row_num, column=7).value = status_record.notes or "-"
            ws2.cell(row=row_num, column=8).value = status_record.admin_notes or "-"
            row_num += 1

        for col_num, _ in enumerate(header2, 1):
            col_letter = get_column_letter(col_num)
            ws2.column_dimensions[col_letter].width = 17

    # Pestaña 3: Detalle de Pausas
    if pause_details:
        ws3 = wb.create_sheet("Detalle de Pausas")
        header3 = [
            "Usuario",
            "Nombre completo",
            "Categoría",
            "Centro",
            "Fecha",
            "Hora Inicio Pausa",
            "Hora Fin Pausa",
            "Tipo de Pausa",
            "Duración",
            "Notas",
        ]
        for col_num, header_text in enumerate(header3, 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        row_num = 2
        for detail in pause_details:
            user = detail["user"]
            pause = detail["pause"]
            ws3.cell(row=row_num, column=1).value = user.username if user else f"ID: {pause.user_id}"
            ws3.cell(row=row_num, column=2).value = user.full_name if user else "-"
            ws3.cell(row=row_num, column=3).value = get_user_category_label(user)
            ws3.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
            ws3.cell(row=row_num, column=5).value = detail["date"].strftime("%d/%m/%Y")
            ws3.cell(row=row_num, column=6).value = pause.pause_start.strftime("%H:%M:%S")
            ws3.cell(row=row_num, column=7).value = pause.pause_end.strftime("%H:%M:%S")
            ws3.cell(row=row_num, column=8).value = pause.pause_type
            ws3.cell(row=row_num, column=9).value = format_duration_from_seconds(detail["duration_seconds"])
            ws3.cell(row=row_num, column=10).value = detail["notes"]
            row_num += 1

        for col_num, _ in enumerate(header3, 1):
            col_letter = get_column_letter(col_num)
            ws3.column_dimensions[col_letter].width = 18

    # ========== PESTAÑA 4: HORAS EXTRAS (condicional) ==========
    # Obtener horas extras SOLO si el filtro está activo
    if 'Horas Extras' in status_filters:
        overtime_entries = OvertimeEntry.query.filter(
            db.or_(
                db.and_(OvertimeEntry.week_start <= fecha, OvertimeEntry.week_end >= fecha),
                OvertimeEntry.week_start == fecha
            )
        ).all()
        add_overtime_sheet_to_workbook(wb, overtime_entries)

    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(temp_path)

    filename = f"{fecha.strftime('%d_%m_%y')}_TimeT.xlsx"
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def handle_daily_pdf_export(req):
    """Maneja exportación PDF diaria con filtros de estado"""
    from fpdf import FPDF

    # Obtener fecha (hoy si no se especifica)
    start_date = req.form.get("start_date")
    end_date = req.form.get("end_date")

    if start_date:
        fecha = datetime.strptime(start_date, "%Y-%m-%d").date()
    else:
        fecha = date.today()

    # Obtener filtros de estado
    status_filters = req.form.getlist('status')
    if not status_filters:
        status_filters = ['Trabajado']

    # Obtener registros de TimeRecord (si "Trabajado" está seleccionado)
    time_records = []
    if 'Trabajado' in status_filters:
        # Eager loading: cargar User en la misma query para evitar N+1
        time_records = (
            TimeRecord.query
            .filter(TimeRecord.date == fecha)
            .options(joinedload(TimeRecord.user), joinedload(TimeRecord.pauses))
            .order_by(TimeRecord.check_in.desc())
            .all()
        )
        time_records = sorted(
            time_records,
            key=lambda r: r.check_in or datetime.combine(r.date, datetime.min.time()),
            reverse=True
        )

    # Obtener registros de EmployeeStatus (Baja, Ausente, Vacaciones)
    employee_statuses = []
    selected_statuses = expand_status_filters([s for s in status_filters if s != 'Trabajado'])
    if selected_statuses:
        employee_statuses = EmployeeStatus.query.filter(
            EmployeeStatus.status.in_(selected_statuses),
            EmployeeStatus.date == fecha
        ).order_by(EmployeeStatus.user_id).all()

    if not time_records and not employee_statuses:
        flash("No hay registros para ese día con los filtros seleccionados.", "warning")
        return redirect(url_for("export.export_excel"))

    # Calcular pausas SOLO si el filtro está activo
    pause_seconds_by_record = {}
    if 'Pausas' in status_filters and time_records:
        pause_seconds_by_record, _ = calculate_pause_data_for_records(time_records)

    # Crear PDF
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Registros del {fecha.strftime('%d/%m/%Y')}", ln=1, align="C")

    # Sección 1: Registros de Fichaje (si hay)
    if time_records:
        pdf.ln(5)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 8, "Registros de Fichaje", ln=1)

        pdf.set_font("Arial", "B", 9)

        # Headers condicionales según filtro de Pausas
        if 'Pausas' in status_filters:
            header = ["Usuario", "Nombre", "Categoria", "Centro", "Entrada", "Salida", "Horas Totales", "Tiempo Pausa", "Horas Efectivas", "Notas", "Notas Admin"]
            col_widths = [20, 28, 22, 25, 18, 18, 22, 22, 22, 35, 35]
        else:
            header = ["Usuario", "Nombre", "Categoria", "Centro", "Entrada", "Salida", "Horas Trabajadas", "Notas", "Notas Admin"]
            col_widths = [25, 35, 28, 30, 20, 20, 28, 50, 50]

        for i, col_name in enumerate(header):
            pdf.cell(col_widths[i], 7, col_name, border=1, align="C")
        pdf.ln()

        pdf.set_font("Arial", "", 8)
        for record in time_records:
            user = record.user
            total_seconds = None
            if record.check_in and record.check_out:
                total_seconds = int((record.check_out - record.check_in).total_seconds())

            # Pausas solo si filtro activo
            pause_seconds = pause_seconds_by_record.get(record.id, 0) if 'Pausas' in status_filters else 0
            effective_seconds = None
            if 'Pausas' in status_filters and total_seconds is not None:
                effective_seconds = max(total_seconds - pause_seconds, 0)

            # Columnas comunes
            pdf.cell(col_widths[0], 6, user.username if user else f"ID:{record.user_id}", border=1)
            pdf.cell(col_widths[1], 6, (user.full_name if user else "-")[:20], border=1)
            pdf.cell(col_widths[2], 6, get_user_category_label(user), border=1)
            pdf.cell(col_widths[3], 6, (user.center.name if user and user.center else "-")[:15], border=1)
            pdf.cell(col_widths[4], 6, record.check_in.strftime("%H:%M") if record.check_in else "-", border=1, align="C")
            pdf.cell(col_widths[5], 6, record.check_out.strftime("%H:%M") if record.check_out else "-", border=1, align="C")

            # Columnas condicionales
            if 'Pausas' in status_filters:
                pdf.cell(col_widths[6], 6, format_duration_from_seconds(total_seconds), border=1, align="C")
                pdf.cell(col_widths[7], 6, format_duration_from_seconds(pause_seconds), border=1, align="C")
                pdf.cell(col_widths[8], 6, format_duration_from_seconds(effective_seconds), border=1, align="C")
                pdf.cell(col_widths[9], 6, (record.notes or "")[:20], border=1)
                pdf.cell(col_widths[10], 6, (record.admin_notes or "")[:20], border=1)
            else:
                pdf.cell(col_widths[6], 6, format_duration_from_seconds(total_seconds), border=1, align="C")
                pdf.cell(col_widths[7], 6, (record.notes or "")[:30], border=1)
                pdf.cell(col_widths[8], 6, (record.admin_notes or "")[:30], border=1)

            pdf.ln()

    # Sección 2: Bajas y Ausencias (si hay)
    if employee_statuses:
        pdf.ln(8)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 8, "Bajas y Ausencias", ln=1)

        pdf.set_font("Arial", "B", 9)
        header2 = ["Usuario", "Nombre", "Categoria", "Centro", "Estado", "Notas", "Notas Admin"]
        col_widths2 = [30, 38, 25, 30, 30, 55, 55]

        for i, col_name in enumerate(header2):
            pdf.cell(col_widths2[i], 7, col_name, border=1, align="C")
        pdf.ln()

        pdf.set_font("Arial", "", 8)
        for status_record in employee_statuses:
            user = User.query.get(status_record.user_id)

            pdf.cell(col_widths2[0], 6, user.username if user else f"ID:{status_record.user_id}", border=1)
            pdf.cell(col_widths2[1], 6, (user.full_name if user else "-")[:25], border=1)
            pdf.cell(col_widths2[2], 6, get_user_category_label(user), border=1)
            pdf.cell(col_widths2[3], 6, (user.center.name if user and user.center else "-")[:20], border=1)
            pdf.cell(col_widths2[4], 6, status_record.status, border=1, align="C")
            pdf.cell(col_widths2[5], 6, (status_record.notes or "")[:35], border=1)
            pdf.cell(col_widths2[6], 6, (status_record.admin_notes or "")[:35], border=1)
            pdf.ln()

    # Sección 3: Horas Extras (condicional)
    if 'Horas Extras' in status_filters:
        # Obtener horas extras de la misma fecha
        overtime_entries = OvertimeEntry.query.filter(
            db.or_(
                db.and_(OvertimeEntry.week_start <= fecha, OvertimeEntry.week_end >= fecha),
                OvertimeEntry.week_start == fecha
            )
        ).all()

        if overtime_entries:
            pdf.ln(8)
            pdf.set_font("Arial", "B", 11)
            pdf.cell(0, 8, "Horas Extras", ln=1)

            pdf.set_font("Arial", "B", 9)
            header3 = ["Usuario", "Nombre", "Semana", "Jornada", "Trabajadas", "Extra/Deficit", "Estado"]
            col_widths3 = [28, 42, 35, 24, 24, 28, 26]

            for i, col_name in enumerate(header3):
                pdf.cell(col_widths3[i], 7, col_name, border=1, align="C")
            pdf.ln()

            pdf.set_font("Arial", "", 8)
            for entry in overtime_entries:
                user = entry.user_rel
                worked_hours = entry.total_worked_seconds / 3600
                contract_hours = entry.contract_seconds / 3600
                overtime_hours = entry.overtime_seconds / 3600

                pdf.cell(col_widths3[0], 6, user.username if user else f"ID:{entry.user_id}", border=1)
                pdf.cell(col_widths3[1], 6, (user.full_name if user else "-")[:25], border=1)
                pdf.cell(col_widths3[2], 6, f"{entry.week_start.strftime('%d/%m')}-{entry.week_end.strftime('%d/%m')}", border=1, align="C")
                pdf.cell(col_widths3[3], 6, f"{contract_hours:.2f}h", border=1, align="C")
                pdf.cell(col_widths3[4], 6, f"{worked_hours:.2f}h", border=1, align="C")
                pdf.cell(col_widths3[5], 6, f"{overtime_hours:+.2f}h", border=1, align="C")
                pdf.cell(col_widths3[6], 6, entry.status[:12], border=1, align="C")
                pdf.ln()

    # Guardar PDF
    fd, temp_path = tempfile.mkstemp(suffix='.pdf')
    os.close(fd)
    pdf.output(temp_path)

    filename = f"{fecha.strftime('%d_%m_%y')}_TimeT.pdf"
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

# ========== EXPORTACIÓN PRINCIPAL CON FILTROS ==========

@export_bp.route("/excel", methods=["GET", "POST"])
@admin_required
def export_excel():
    if request.method == "POST":
        # Detectar botones de Excel/PDF Diario
        if 'export_daily_excel' in request.form:
            return handle_daily_excel_export(request)
        if 'export_daily_pdf' in request.form:
            return handle_daily_pdf_export(request)

        # Detectar qué botón se ha pulsado realmente (el último en el array de botones recibidos)
        botones = [k for k in request.form.keys() if k.startswith('excel_')]
        boton_pulsado = botones[-1] if botones else None

        # LOG de depuracion para ver los valores recibidos
        logger.debug('--- FILTROS RECIBIDOS ---')
        logger.debug(f'Botones recibidos: {botones}')
        logger.debug(f'Boton realmente pulsado: {boton_pulsado}')
        logger.debug(f"centro1: {request.form.get('centro1')}")
        logger.debug(f"usuario1: {request.form.get('usuario1')}")
        logger.debug(f"centro2: {request.form.get('centro2')}")
        logger.debug(f"categoria2: {request.form.get('categoria2')}")
        logger.debug(f"centro3: {request.form.get('centro3')}")
        logger.debug(f"horas3: {request.form.get('horas3')}")
        logger.debug(f"centro4: {request.form.get('centro4')}")
        logger.debug(f"usuario4: {request.form.get('usuario4')}")
        logger.debug(f"categoria4: {request.form.get('categoria4')}")
        logger.debug(f"horas4: {request.form.get('horas4')}")
        logger.debug(f"start_date: {request.form.get('start_date')}")
        logger.debug(f"end_date: {request.form.get('end_date')}")

        # Inicializar filtros
        centro = user_id = categoria = weekly_hours = None

        # Usar los campos correctos según el botón realmente pulsado
        if boton_pulsado == "excel_centro_usuario":
            centro = request.form.get("centro1")
            user_id = request.form.get("usuario1")
        elif boton_pulsado == "excel_centro_categoria":
            centro = request.form.get("centro2")
            categoria = request.form.get("categoria2")
        elif boton_pulsado == "excel_centro_horas":
            centro = request.form.get("centro3")
            weekly_hours = request.form.get("horas3")
        elif boton_pulsado == "excel_solo_centro":
            centro = request.form.get("centro4")
        elif boton_pulsado == "excel_solo_usuario":
            user_id = request.form.get("usuario4")
        elif boton_pulsado == "excel_solo_categoria":
            categoria = request.form.get("categoria4")
        elif boton_pulsado == "excel_solo_horas":
            weekly_hours = request.form.get("horas4")
        else:
            # Compatibilidad con los filtros antiguos
            centro = request.form.get("centro")
            user_id = request.form.get("user_id")
            categoria = request.form.get("categoria")
            weekly_hours = request.form.get("weekly_hours") or request.form.get("jornada")

        categoria_id_filter, categoria_none_filter = resolve_category_filter(categoria)

        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")

        # Validación fechas
        try:
            if start_date:
                start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            else:
                start_date = date.today()
            if end_date:
                end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            else:
                end_date = date.today()
            if end_date < start_date:
                flash("La fecha de fin no puede ser anterior a la fecha de inicio.", "danger")
                return redirect(url_for("export.export_excel"))
        except ValueError:
            flash("Formato de fecha inválido. Use YYYY-MM-DD.", "danger")
            return redirect(url_for("export.export_excel"))

        # Obtener filtros de estado
        status_filters = list(dict.fromkeys(request.form.getlist('status')))
        if not status_filters:
            status_filters = ['Trabajado']  # Default solo trabajado

        categoria_id_filter, categoria_none_filter = resolve_category_filter(categoria)

        logger.debug('Valores usados para filtrar:')
        logger.debug(f'centro: {centro}')
        logger.debug(f'user_id: {user_id}')
        logger.debug(f'categoria: {categoria}')
        logger.debug(f'weekly_hours: {weekly_hours}')
        logger.debug(f'status_filters: {status_filters}')
        logger.debug('--------------------------')

        # ========== OBTENER REGISTROS DE TIMERECORD (si "Trabajado" está seleccionado) ==========
        time_records = []
        if 'Trabajado' in status_filters:
            query = TimeRecord.query.join(User, TimeRecord.user_id == User.id).filter(
                TimeRecord.date >= start_date,
                TimeRecord.date <= end_date
            )

            # Aplicar filtros
            if centro:
                query = query.filter(User.center_id == centro)
            if user_id:
                query = query.filter(TimeRecord.user_id == user_id)
            if categoria_id_filter:
                query = query.filter(User.category_id == categoria_id_filter)
            elif categoria_none_filter:
                query = query.filter(User.category_id.is_(None))
            if weekly_hours:
                try:
                    wh = int(weekly_hours)
                    query = query.filter(User.weekly_hours.isnot(None))
                    query = query.filter(db.cast(User.weekly_hours, db.Integer) == wh)
                except ValueError:
                    flash("La jornada debe ser numérica.", "danger")
                    return redirect(url_for("export.export_excel"))

            time_records = query.order_by(TimeRecord.user_id, TimeRecord.date).all()

        # ========== OBTENER REGISTROS DE EMPLOYEESTATUS (Baja, Ausente, Vacaciones) ==========
        employee_statuses = []
        selected_statuses = expand_status_filters([s for s in status_filters if s != 'Trabajado'])
        if selected_statuses:
            status_query = EmployeeStatus.query.join(User, EmployeeStatus.user_id == User.id).filter(
                EmployeeStatus.status.in_(selected_statuses),
                EmployeeStatus.date >= start_date,
                EmployeeStatus.date <= end_date
            )

            # Aplicar los mismos filtros
            if centro:
                status_query = status_query.filter(User.center_id == centro)
            if user_id:
                status_query = status_query.filter(EmployeeStatus.user_id == user_id)
            if categoria_id_filter:
                status_query = status_query.filter(User.category_id == categoria_id_filter)
            elif categoria_none_filter:
                status_query = status_query.filter(User.category_id.is_(None))
            if weekly_hours:
                try:
                    wh = int(weekly_hours)
                    status_query = status_query.filter(User.weekly_hours.isnot(None))
                    status_query = status_query.filter(db.cast(User.weekly_hours, db.Integer) == wh)
                except ValueError:
                    pass  # Ya se validó arriba

            employee_statuses = status_query.order_by(EmployeeStatus.user_id, EmployeeStatus.date).all()

        # Verificar que hay al menos algún registro
        if not time_records and not employee_statuses:
            flash("No hay registros para el período y filtros seleccionados.", "warning")
            return redirect(url_for("export.export_excel"))

        # Calcular pausas si el filtro está activo
        pause_seconds_by_record = {}
        pause_details = []
        if 'Pausas' in status_filters and time_records:
            pause_seconds_by_record, pause_details = calculate_pause_data_for_records(time_records)

        # ========== GENERAR EXCEL CON 3 PESTAÑAS ==========

        wb = openpyxl.Workbook()

        # ========== PESTAÑA 1: REGISTROS DE FICHAJE ==========
        if time_records:
            ws1 = wb.active
            ws1.title = "Registros de Fichaje"

            # Header condicional según filtro de Pausas
            if 'Pausas' in status_filters:
                header1 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Entrada", "Salida", "Horas Totales", "Tiempo de Pausa", "Horas Efectivas", "Notas", "Notas Admin", "Modificado Por", "Última Actualización"]
            else:
                header1 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Entrada", "Salida", "Horas Trabajadas", "Notas", "Notas Admin", "Modificado Por", "Última Actualización"]

            for col_num, header_text in enumerate(header1, 1):
                cell = ws1.cell(row=1, column=col_num)
                cell.value = header_text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            row_num = 2
            for record in time_records:
                user = User.query.get(record.user_id)
                modified_by = User.query.get(record.modified_by) if record.modified_by else None

                # Calcular horas totales
                total_seconds = None
                if record.check_in and record.check_out:
                    time_diff = record.check_out - record.check_in
                    total_seconds = int(time_diff.total_seconds())

                # Pausas solo si filtro activo
                pause_seconds = pause_seconds_by_record.get(record.id, 0) if 'Pausas' in status_filters else 0
                effective_seconds = None
                if 'Pausas' in status_filters and total_seconds is not None:
                    effective_seconds = max(total_seconds - pause_seconds, 0)

                # Escribir columnas comunes (1-7)
                ws1.cell(row=row_num, column=1).value = user.username if user else f"ID: {record.user_id}"
                ws1.cell(row=row_num, column=2).value = user.full_name if user else "-"
                ws1.cell(row=row_num, column=3).value = get_user_category_label(user)
                ws1.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
                ws1.cell(row=row_num, column=5).value = record.date.strftime("%d/%m/%Y")
                ws1.cell(row=row_num, column=6).value = record.check_in.strftime("%H:%M:%S") if record.check_in else "-"
                ws1.cell(row=row_num, column=7).value = record.check_out.strftime("%H:%M:%S") if record.check_out else "-"

                # Columnas condicionales según filtro
                if 'Pausas' in status_filters:
                    ws1.cell(row=row_num, column=8).value = format_duration_from_seconds(total_seconds)
                    ws1.cell(row=row_num, column=9).value = format_duration_from_seconds(pause_seconds)
                    ws1.cell(row=row_num, column=10).value = format_duration_from_seconds(effective_seconds)
                    ws1.cell(row=row_num, column=11).value = record.notes
                    ws1.cell(row=row_num, column=12).value = record.admin_notes
                    ws1.cell(row=row_num, column=13).value = modified_by.username if modified_by else "-"
                    ws1.cell(row=row_num, column=14).value = record.updated_at.strftime("%d/%m/%Y %H:%M:%S")
                else:
                    ws1.cell(row=row_num, column=8).value = format_duration_from_seconds(total_seconds)
                    ws1.cell(row=row_num, column=9).value = record.notes
                    ws1.cell(row=row_num, column=10).value = record.admin_notes
                    ws1.cell(row=row_num, column=11).value = modified_by.username if modified_by else "-"
                    ws1.cell(row=row_num, column=12).value = record.updated_at.strftime("%d/%m/%Y %H:%M:%S")

                row_num += 1

            for col_num, _ in enumerate(header1, 1):
                col_letter = get_column_letter(col_num)
                ws1.column_dimensions[col_letter].width = 17
        else:
            # Si no hay registros de fichaje, eliminar la primera pestaña
            wb.remove(wb.active)

        # ========== PESTAÑA 2: BAJAS Y AUSENCIAS ==========
        if employee_statuses:
            ws2 = wb.create_sheet("Bajas y Ausencias")

            header2 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Estado", "Notas", "Notas Admin"]
            for col_num, header_text in enumerate(header2, 1):
                cell = ws2.cell(row=1, column=col_num)
                cell.value = header_text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            row_num = 2
            for status_record in employee_statuses:
                user = User.query.get(status_record.user_id)

                ws2.cell(row=row_num, column=1).value = user.username if user else f"ID: {status_record.user_id}"
                ws2.cell(row=row_num, column=2).value = user.full_name if user else "-"
                ws2.cell(row=row_num, column=3).value = get_user_category_label(user)
                ws2.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
                ws2.cell(row=row_num, column=5).value = status_record.date.strftime("%d/%m/%Y")
                ws2.cell(row=row_num, column=6).value = status_record.status
                ws2.cell(row=row_num, column=7).value = status_record.notes or "-"
                ws2.cell(row=row_num, column=8).value = status_record.admin_notes or "-"
                row_num += 1

            for col_num, _ in enumerate(header2, 1):
                col_letter = get_column_letter(col_num)
                ws2.column_dimensions[col_letter].width = 17

        # ========== PESTAÑA: DETALLE DE PAUSAS ==========
        if pause_details:
            ws_pausas = wb.create_sheet("Detalle de Pausas")
            header_pausas = [
                "Usuario", "Nombre completo", "Categoría", "Centro", "Fecha",
                "Hora Inicio Pausa", "Hora Fin Pausa", "Tipo de Pausa",
                "Duración", "Notas"
            ]

            for col_num, header_text in enumerate(header_pausas, 1):
                cell = ws_pausas.cell(row=1, column=col_num)
                cell.value = header_text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            row_num = 2
            for detail in pause_details:
                user = detail["user"]
                pause = detail["pause"]
                ws_pausas.cell(row=row_num, column=1).value = user.username if user else f"ID: {pause.user_id}"
                ws_pausas.cell(row=row_num, column=2).value = user.full_name if user else "-"
                ws_pausas.cell(row=row_num, column=3).value = get_user_category_label(user)
                ws_pausas.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
                ws_pausas.cell(row=row_num, column=5).value = detail["date"].strftime("%d/%m/%Y")
                ws_pausas.cell(row=row_num, column=6).value = pause.pause_start.strftime("%H:%M:%S")
                ws_pausas.cell(row=row_num, column=7).value = pause.pause_end.strftime("%H:%M:%S")
                ws_pausas.cell(row=row_num, column=8).value = pause.pause_type
                ws_pausas.cell(row=row_num, column=9).value = format_duration_from_seconds(detail["duration_seconds"])
                ws_pausas.cell(row=row_num, column=10).value = detail["notes"]
                row_num += 1

            for col_num, _ in enumerate(header_pausas, 1):
                col_letter = get_column_letter(col_num)
                ws_pausas.column_dimensions[col_letter].width = 18

        # ========== PESTAÑA 3: RESUMEN CONSOLIDADO ==========
        ws3 = wb.create_sheet("Resumen Consolidado")

        header3 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Estado", "Entrada", "Salida", "Horas", "Notas", "Notas Admin"]
        for col_num, header_text in enumerate(header3, 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Combinar ambos tipos de registros en una lista unificada
        consolidated_records = []

        # Agregar TimeRecords
        for record in time_records:
            user = User.query.get(record.user_id)
            hours_worked = ""
            if record.check_in and record.check_out:
                time_diff = record.check_out - record.check_in
                hours = time_diff.total_seconds() / 3600
                hours_worked = f"{hours:.2f}"

            consolidated_records.append({
                'user_id': record.user_id,
                'username': user.username if user else f"ID: {record.user_id}",
                'full_name': user.full_name if user else "-",
                'categoria': get_user_category_label(user),
                'centro': user.center.name if user and user.center else "-",
                'date': record.date,
                'estado': "Trabajado",
                'entrada': record.check_in.strftime("%H:%M:%S") if record.check_in else "-",
                'salida': record.check_out.strftime("%H:%M:%S") if record.check_out else "-",
                'horas': hours_worked,
                'notas': record.notes or "-",
                'admin_notes': record.admin_notes or "-"
            })

        # Agregar EmployeeStatus
        for status_record in employee_statuses:
            user = User.query.get(status_record.user_id)
            consolidated_records.append({
                'user_id': status_record.user_id,
                'username': user.username if user else f"ID: {status_record.user_id}",
                'full_name': user.full_name if user else "-",
                'categoria': get_user_category_label(user),
                'centro': user.center.name if user and user.center else "-",
                'date': status_record.date,
                'estado': status_record.status,
                'entrada': "-",
                'salida': "-",
                'horas': "-",
                'notas': status_record.notes or "-",
                'admin_notes': status_record.admin_notes or "-"
            })

        # Ordenar por usuario y fecha
        consolidated_records.sort(key=lambda x: (x['user_id'], x['date']))

        # Escribir registros consolidados
        row_num = 2
        for rec in consolidated_records:
            ws3.cell(row=row_num, column=1).value = rec['username']
            ws3.cell(row=row_num, column=2).value = rec['full_name']
            ws3.cell(row=row_num, column=3).value = rec['categoria']
            ws3.cell(row=row_num, column=4).value = rec['centro']
            ws3.cell(row=row_num, column=5).value = rec['date'].strftime("%d/%m/%Y")
            ws3.cell(row=row_num, column=6).value = rec['estado']
            ws3.cell(row=row_num, column=7).value = rec['entrada']
            ws3.cell(row=row_num, column=8).value = rec['salida']
            ws3.cell(row=row_num, column=9).value = rec['horas']
            ws3.cell(row=row_num, column=10).value = rec['notas']
            ws3.cell(row=row_num, column=11).value = rec['admin_notes']
            row_num += 1

        for col_num, _ in enumerate(header3, 1):
            col_letter = get_column_letter(col_num)
            ws3.column_dimensions[col_letter].width = 17

        # ========== PESTAÑA 4: HORAS EXTRAS (condicional) ==========
        # Añadir pestaña de horas extras SOLO si el filtro está activo
        if 'Horas Extras' in status_filters:
            # Filtrar horas extras por rango de fechas
            overtime_query = OvertimeEntry.query.filter(
                db.or_(
                    db.and_(OvertimeEntry.week_start <= end_date, OvertimeEntry.week_end >= start_date),
                    db.and_(OvertimeEntry.week_start >= start_date, OvertimeEntry.week_start <= end_date)
                )
            )

            # Aplicar filtros de usuario si existen
            if user_id:
                overtime_query = overtime_query.filter(OvertimeEntry.user_id == user_id)
            elif categoria_id_filter:
                overtime_query = overtime_query.join(User).filter(User.category_id == categoria_id_filter)
            elif categoria_none_filter:
                overtime_query = overtime_query.join(User).filter(User.category_id.is_(None))

            if centro:
                overtime_query = overtime_query.join(User, OvertimeEntry.user_id == User.id).filter(User.center_id == centro)

            overtime_entries = overtime_query.all()
            add_overtime_sheet_to_workbook(wb, overtime_entries)

        fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(temp_path)

        # Generar nombre con formato dd_mm_yy_TimeT
        filename = f"{end_date.strftime('%d_%m_%y')}_TimeT.xlsx"
        return send_file(
            temp_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # GET
    from routes.admin import get_admin_centro, get_centros_dinamicos, get_categorias_disponibles
    centro_admin = get_admin_centro()
    q = User.query.filter_by(is_active=True)
    if centro_admin:
        q = q.filter(User.center_id == centro_admin)
    users = q.order_by(User.username).all()

    # Calcular lista de horas únicas y ordenadas ascendentemente para el desplegable "horas4"
    horas_sorted = sorted({u.weekly_hours for u in users if u.weekly_hours is not None})

    # Obtener centros y categorías dinámicos
    centros = get_centros_dinamicos()
    categorias = get_categorias_disponibles()

    today = date.today().strftime('%Y-%m-%d')
    return render_template("export_excel.html", users=users, today=today, centro_admin=centro_admin, horas_sorted=horas_sorted, centros=centros, categorias=categorias)

# ========== EXCEL MENSUAL CON SUMAS SEMANALES ==========

@export_bp.route("/excel_monthly", methods=["GET", "POST"])
@admin_required
def export_excel_monthly():
    if request.method == "POST":
        # Obtener filtros de estado
        status_filters = request.form.getlist('status')
        if not status_filters:
            status_filters = ['Trabajado']

        # Reutilizar la misma lógica de filtros que la función principal
        botones = [k for k in request.form.keys() if k.startswith('excel_')]
        boton_pulsado = botones[-1] if botones else None

        centro = user_id = categoria = weekly_hours = None
        if boton_pulsado == "excel_centro_usuario":
            centro = request.form.get("centro1")
            user_id = request.form.get("usuario1")
        elif boton_pulsado == "excel_centro_categoria":
            centro = request.form.get("centro2")
            categoria = request.form.get("categoria2")
        elif boton_pulsado == "excel_centro_horas":
            centro = request.form.get("centro3")
            weekly_hours = request.form.get("horas3")
        elif boton_pulsado == "excel_solo_centro":
            centro = request.form.get("centro4")
        elif boton_pulsado == "excel_solo_usuario":
            user_id = request.form.get("usuario4")
        elif boton_pulsado == "excel_solo_categoria":
            categoria = request.form.get("categoria4")
        elif boton_pulsado == "excel_solo_horas":
            weekly_hours = request.form.get("horas4")
        else:
            centro = request.form.get("centro")
            user_id = request.form.get("user_id")
            categoria = request.form.get("categoria")
            weekly_hours = request.form.get("weekly_hours") or request.form.get("jornada")

        categoria_id_filter, categoria_none_filter = resolve_category_filter(categoria)

        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")

        # Validación fechas
        try:
            if start_date:
                start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            else:
                start_date = date.today()
            if end_date:
                end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            else:
                end_date = date.today()
            if end_date < start_date:
                flash("La fecha de fin no puede ser anterior a la fecha de inicio.", "danger")
                return redirect(url_for("export.export_excel_monthly"))
        except ValueError:
            flash("Formato de fecha inválido. Use YYYY-MM-DD.", "danger")
            return redirect(url_for("export.export_excel_monthly"))

        # Consultar TimeRecord si "Trabajado" está seleccionado
        records = []
        if 'Trabajado' in status_filters:
            query = TimeRecord.query.join(User, TimeRecord.user_id == User.id).filter(
                TimeRecord.date >= start_date,
                TimeRecord.date <= end_date
            )

            if centro:
                query = query.filter(User.center_id == centro)
            if user_id:
                query = query.filter(TimeRecord.user_id == user_id)
            if categoria_id_filter:
                query = query.filter(User.category_id == categoria_id_filter)
            elif categoria_none_filter:
                query = query.filter(User.category_id.is_(None))
            if weekly_hours:
                try:
                    wh = int(weekly_hours)
                    query = query.filter(User.weekly_hours.isnot(None))
                    query = query.filter(db.cast(User.weekly_hours, db.Integer) == wh)
                except ValueError:
                    flash("La jornada debe ser numérica.", "danger")
                    return redirect(url_for("export.export_excel_monthly"))

            records = query.order_by(TimeRecord.user_id, TimeRecord.date).all()

        # Consultar EmployeeStatus si otros estados están seleccionados
        employee_statuses = []
        selected_statuses = expand_status_filters([s for s in status_filters if s != 'Trabajado'])
        if selected_statuses:
            status_query = EmployeeStatus.query.join(User, EmployeeStatus.user_id == User.id).filter(
                EmployeeStatus.status.in_(selected_statuses),
                EmployeeStatus.date >= start_date,
                EmployeeStatus.date <= end_date
            )

            if centro:
                status_query = status_query.filter(User.center_id == centro)
            if user_id:
                status_query = status_query.filter(EmployeeStatus.user_id == user_id)
            if categoria_id_filter:
                status_query = status_query.filter(User.category_id == categoria_id_filter)
            elif categoria_none_filter:
                status_query = status_query.filter(User.category_id.is_(None))
            if weekly_hours:
                try:
                    wh = int(weekly_hours)
                    status_query = status_query.filter(User.weekly_hours.isnot(None))
                    status_query = status_query.filter(db.cast(User.weekly_hours, db.Integer) == wh)
                except ValueError:
                    pass

            employee_statuses = status_query.order_by(EmployeeStatus.user_id, EmployeeStatus.date).all()

        if not records and not employee_statuses:
            flash("No hay registros para el período y filtros seleccionados.", "warning")
            return redirect(url_for("export.export_excel_monthly"))

        # Calcular pausas si el filtro está activo
        pause_seconds_by_record = {}
        pause_details = []
        if 'Pausas' in status_filters and records:
            pause_seconds_by_record, pause_details = calculate_pause_data_for_records(records)

        # Generar Excel con pestañas condicionales (hasta 5 pestañas)
        wb = openpyxl.Workbook()

        # ===== PESTAÑA 1: REGISTROS DE FICHAJE (con sumas semanales) =====
        if records:
            ws1 = wb.active
            ws1.title = "Registros de Fichaje"

            # Usar la misma lógica de sumas semanales
            def get_week_start(date_obj):
                return date_obj - timedelta(days=date_obj.weekday())

            weekly_data = defaultdict(lambda: defaultdict(list))
            weekly_totals = defaultdict(lambda: defaultdict(float))

            for record in records:
                week_start = get_week_start(record.date)
                weekly_data[record.user_id][week_start].append(record)

                if record.check_in and record.check_out:
                    time_diff = record.check_out - record.check_in
                    hours = time_diff.total_seconds() / 3600
                    weekly_totals[record.user_id][week_start] += hours

            # Header condicional según filtro de Pausas
            if 'Pausas' in status_filters:
                header1 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Horas Semanales", "Fecha", "Entrada", "Salida", "Horas Totales", "Tiempo de Pausa", "Horas Efectivas", "Diferencia Horas", "Notas", "Notas Admin", "Modificado Por", "Última Actualización"]
            else:
                header1 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Horas Semanales", "Fecha", "Entrada", "Salida", "Horas Trabajadas", "Diferencia Horas", "Notas", "Notas Admin", "Modificado Por", "Última Actualización"]

            for col_num, header_text in enumerate(header1, 1):
                cell = ws1.cell(row=1, column=col_num)
                cell.value = header_text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            row_num = 2

            for user_id in sorted(weekly_data.keys()):
                user = User.query.get(user_id)

                for week_start in sorted(weekly_data[user_id].keys()):
                    week_end = week_start + timedelta(days=6)
                    total_hours = weekly_totals[user_id][week_start]

                    # Calcular total de pausas de la semana si filtro activo
                    total_pause_hours = 0
                    if 'Pausas' in status_filters:
                        for record in weekly_data[user_id][week_start]:
                            pause_seconds = pause_seconds_by_record.get(record.id, 0)
                            total_pause_hours += pause_seconds / 3600

                    # Fila de total semanal
                    cell = ws1.cell(row=row_num, column=1)
                    cell.value = f"TOTAL SEMANA ({week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')})"
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                    cell = ws1.cell(row=row_num, column=2)
                    cell.value = user.full_name if user else "-"
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                    # Columnas 3 y 4: "-"
                    for col in range(3, 5):
                        cell = ws1.cell(row=row_num, column=col)
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                    # Columna 5: Mostrar las horas semanales contractuales
                    cell = ws1.cell(row=row_num, column=5)
                    cell.value = user.weekly_hours if user and user.weekly_hours else "-"
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                    # Columnas 6, 7, 8: "-"
                    for col in range(6, 9):
                        cell = ws1.cell(row=row_num, column=col)
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                    # Columna 9: Horas totales trabajadas
                    cell = ws1.cell(row=row_num, column=9)
                    cell.value = f"{total_hours:.2f}"
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                    # Columnas condicionales según filtro de Pausas
                    if 'Pausas' in status_filters:
                        # Columna 10: Tiempo de pausa
                        cell = ws1.cell(row=row_num, column=10)
                        cell.value = f"{total_pause_hours:.2f}"
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                        # Columna 11: Horas efectivas
                        effective_hours = total_hours - total_pause_hours
                        cell = ws1.cell(row=row_num, column=11)
                        cell.value = f"{effective_hours:.2f}"
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                        # Columna 12: Diferencia
                        weekly_hours_contract = user.weekly_hours if user and user.weekly_hours else 0
                        difference = weekly_hours_contract - total_hours
                        cell = ws1.cell(row=row_num, column=12)
                        cell.value = f"{difference:.2f}"
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                        # Columnas 13, 14, 15, 16: "-"
                        for col in range(13, 17):
                            cell = ws1.cell(row=row_num, column=col)
                            cell.value = "-"
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    else:
                        # Columna 10: Diferencia (sin pausas)
                        weekly_hours_contract = user.weekly_hours if user and user.weekly_hours else 0
                        difference = weekly_hours_contract - total_hours
                        cell = ws1.cell(row=row_num, column=10)
                        cell.value = f"{difference:.2f}"
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                        # Columnas 11, 12, 13, 14: "-"
                        for col in range(11, 15):
                            cell = ws1.cell(row=row_num, column=col)
                            cell.value = "-"
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                    row_num += 1

                    # Registros individuales
                    for record in weekly_data[user_id][week_start]:
                        modified_by = User.query.get(record.modified_by) if record.modified_by else None
                        total_seconds = 0
                        if record.check_in and record.check_out:
                            time_diff = record.check_out - record.check_in
                            total_seconds = time_diff.total_seconds()

                        # Calcular pausas si filtro activo
                        pause_seconds = pause_seconds_by_record.get(record.id, 0) if 'Pausas' in status_filters else 0
                        effective_seconds = max(total_seconds - pause_seconds, 0) if 'Pausas' in status_filters else 0

                        # Escribir columnas comunes (1-8)
                        ws1.cell(row=row_num, column=1).value = user.username if user else f"ID: {record.user_id}"
                        ws1.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')

                        ws1.cell(row=row_num, column=2).value = user.full_name if user else "-"
                        ws1.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')

                        ws1.cell(row=row_num, column=3).value = get_user_category_label(user)
                        ws1.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')

                        ws1.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
                        ws1.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')

                        ws1.cell(row=row_num, column=5).value = user.weekly_hours if user and user.weekly_hours else "-"
                        ws1.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')

                        ws1.cell(row=row_num, column=6).value = record.date.strftime("%d/%m/%Y")
                        ws1.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')

                        ws1.cell(row=row_num, column=7).value = record.check_in.strftime("%H:%M:%S") if record.check_in else "-"
                        ws1.cell(row=row_num, column=7).alignment = Alignment(horizontal='center')

                        ws1.cell(row=row_num, column=8).value = record.check_out.strftime("%H:%M:%S") if record.check_out else "-"
                        ws1.cell(row=row_num, column=8).alignment = Alignment(horizontal='center')

                        # Columnas condicionales según filtro de Pausas
                        if 'Pausas' in status_filters:
                            ws1.cell(row=row_num, column=9).value = f"{total_seconds / 3600:.2f}" if total_seconds else "-"
                            ws1.cell(row=row_num, column=9).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=10).value = f"{pause_seconds / 3600:.2f}" if pause_seconds else "0.00"
                            ws1.cell(row=row_num, column=10).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=11).value = f"{effective_seconds / 3600:.2f}" if effective_seconds else "-"
                            ws1.cell(row=row_num, column=11).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=12).value = "-"
                            ws1.cell(row=row_num, column=12).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=13).value = record.notes
                            ws1.cell(row=row_num, column=13).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=14).value = record.admin_notes
                            ws1.cell(row=row_num, column=14).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=15).value = modified_by.username if modified_by else "-"
                            ws1.cell(row=row_num, column=15).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=16).value = record.updated_at.strftime("%d/%m/%Y %H:%M:%S")
                            ws1.cell(row=row_num, column=16).alignment = Alignment(horizontal='center')
                        else:
                            ws1.cell(row=row_num, column=9).value = f"{total_seconds / 3600:.2f}" if total_seconds else "-"
                            ws1.cell(row=row_num, column=9).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=10).value = "-"
                            ws1.cell(row=row_num, column=10).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=11).value = record.notes
                            ws1.cell(row=row_num, column=11).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=12).value = record.admin_notes
                            ws1.cell(row=row_num, column=12).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=13).value = modified_by.username if modified_by else "-"
                            ws1.cell(row=row_num, column=13).alignment = Alignment(horizontal='center')

                            ws1.cell(row=row_num, column=14).value = record.updated_at.strftime("%d/%m/%Y %H:%M:%S")
                            ws1.cell(row=row_num, column=14).alignment = Alignment(horizontal='center')
                        row_num += 1

                    row_num += 1

            for col_num, _ in enumerate(header1, 1):
                col_letter = get_column_letter(col_num)
                ws1.column_dimensions[col_letter].width = 17
        else:
            wb.remove(wb.active)

        # ===== PESTAÑA 2: BAJAS Y AUSENCIAS =====
        if employee_statuses:
            ws2 = wb.create_sheet("Bajas y Ausencias")
            header2 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Estado", "Notas", "Notas Admin"]
            for col_num, header_text in enumerate(header2, 1):
                cell = ws2.cell(row=1, column=col_num)
                cell.value = header_text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            row_num = 2
            for status_record in employee_statuses:
                user = User.query.get(status_record.user_id)
                ws2.cell(row=row_num, column=1).value = user.username if user else f"ID: {status_record.user_id}"
                ws2.cell(row=row_num, column=2).value = user.full_name if user else "-"
                ws2.cell(row=row_num, column=3).value = get_user_category_label(user)
                ws2.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
                ws2.cell(row=row_num, column=5).value = status_record.date.strftime("%d/%m/%Y")
                ws2.cell(row=row_num, column=6).value = status_record.status
                ws2.cell(row=row_num, column=7).value = status_record.notes or "-"
                ws2.cell(row=row_num, column=8).value = status_record.admin_notes or "-"
                row_num += 1

            for col_num, _ in enumerate(header2, 1):
                col_letter = get_column_letter(col_num)
                ws2.column_dimensions[col_letter].width = 17

        # ===== PESTAÑA 3: DETALLE DE PAUSAS (condicional) =====
        if pause_details:
            ws_pausas = wb.create_sheet("Detalle de Pausas")
            header_pausas = [
                "Usuario", "Nombre completo", "Categoría", "Centro", "Fecha",
                "Hora Inicio Pausa", "Hora Fin Pausa", "Tipo de Pausa",
                "Duración", "Notas"
            ]

            for col_num, header_text in enumerate(header_pausas, 1):
                cell = ws_pausas.cell(row=1, column=col_num)
                cell.value = header_text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            row_num = 2
            for detail in pause_details:
                user = detail["user"]
                pause = detail["pause"]
                ws_pausas.cell(row=row_num, column=1).value = user.username if user else f"ID: {pause.user_id}"
                ws_pausas.cell(row=row_num, column=2).value = user.full_name if user else "-"
                ws_pausas.cell(row=row_num, column=3).value = get_user_category_label(user)
                ws_pausas.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
                ws_pausas.cell(row=row_num, column=5).value = detail["date"].strftime("%d/%m/%Y")
                ws_pausas.cell(row=row_num, column=6).value = pause.pause_start.strftime("%H:%M:%S")
                ws_pausas.cell(row=row_num, column=7).value = pause.pause_end.strftime("%H:%M:%S")
                ws_pausas.cell(row=row_num, column=8).value = pause.pause_type
                ws_pausas.cell(row=row_num, column=9).value = format_duration_from_seconds(detail["duration_seconds"])
                ws_pausas.cell(row=row_num, column=10).value = detail["notes"]
                row_num += 1

            for col_num, _ in enumerate(header_pausas, 1):
                col_letter = get_column_letter(col_num)
                ws_pausas.column_dimensions[col_letter].width = 18

        # ===== PESTAÑA 4: RESUMEN CONSOLIDADO =====
        ws3 = wb.create_sheet("Resumen Consolidado")
        header3 = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Estado", "Entrada", "Salida", "Horas", "Notas", "Notas Admin"]
        for col_num, header_text in enumerate(header3, 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        consolidated_records = []
        for record in records:
            user = User.query.get(record.user_id)
            hours_worked = ""
            if record.check_in and record.check_out:
                time_diff = record.check_out - record.check_in
                hours = time_diff.total_seconds() / 3600
                hours_worked = f"{hours:.2f}"
            consolidated_records.append({
                'user_id': record.user_id,
                'username': user.username if user else f"ID: {record.user_id}",
                'full_name': user.full_name if user else "-",
                'categoria': get_user_category_label(user),
                'centro': user.center.name if user and user.center else "-",
                'date': record.date,
                'estado': "Trabajado",
                'entrada': record.check_in.strftime("%H:%M:%S") if record.check_in else "-",
                'salida': record.check_out.strftime("%H:%M:%S") if record.check_out else "-",
                'horas': hours_worked,
                'notas': record.notes or "-",
                'admin_notes': record.admin_notes or "-"
            })

        for status_record in employee_statuses:
            user = User.query.get(status_record.user_id)
            consolidated_records.append({
                'user_id': status_record.user_id,
                'username': user.username if user else f"ID: {status_record.user_id}",
                'full_name': user.full_name if user else "-",
                'categoria': get_user_category_label(user),
                'centro': user.center.name if user and user.center else "-",
                'date': status_record.date,
                'estado': status_record.status,
                'entrada': "-",
                'salida': "-",
                'horas': "-",
                'notas': status_record.notes or "-",
                'admin_notes': status_record.admin_notes or "-"
            })

        # Ordenar por usuario y fecha
        consolidated_records.sort(key=lambda x: (x['user_id'], x['date']))

        row_num = 2
        for rec in consolidated_records:
            ws3.cell(row=row_num, column=1).value = rec['username']
            ws3.cell(row=row_num, column=2).value = rec['full_name']
            ws3.cell(row=row_num, column=3).value = rec['categoria']
            ws3.cell(row=row_num, column=4).value = rec['centro']
            ws3.cell(row=row_num, column=5).value = rec['date'].strftime("%d/%m/%Y")
            ws3.cell(row=row_num, column=6).value = rec['estado']
            ws3.cell(row=row_num, column=7).value = rec['entrada']
            ws3.cell(row=row_num, column=8).value = rec['salida']
            ws3.cell(row=row_num, column=9).value = rec['horas']
            ws3.cell(row=row_num, column=10).value = rec['notas']
            ws3.cell(row=row_num, column=11).value = rec['admin_notes']
            row_num += 1

        for col_num, _ in enumerate(header3, 1):
            col_letter = get_column_letter(col_num)
            ws3.column_dimensions[col_letter].width = 17

        # ========== PESTAÑA 5: HORAS EXTRAS (condicional) ==========
        # Añadir pestaña de horas extras SOLO si el filtro está activo
        if 'Horas Extras' in status_filters:
            # Filtrar horas extras del mes seleccionado
            overtime_entries = OvertimeEntry.query.filter(
                db.or_(
                    db.and_(OvertimeEntry.week_start >= start_date, OvertimeEntry.week_start <= end_date),
                    db.and_(OvertimeEntry.week_end >= start_date, OvertimeEntry.week_end <= end_date)
                )
            ).all()
            add_overtime_sheet_to_workbook(wb, overtime_entries)

        fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(temp_path)

        # Generar nombre con formato dd_mm_yy_TimeT
        from datetime import date as dt_date
        today = dt_date.today()
        filename = f"{today.strftime('%d_%m_%y')}_TimeT.xlsx"

        return send_file(
            temp_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # GET - usar el mismo template
    from routes.admin import get_admin_centro, get_centros_dinamicos, get_categorias_disponibles
    centro_admin = get_admin_centro()
    q = User.query.filter_by(is_active=True)
    if centro_admin:
        q = q.filter(User.center_id == centro_admin)
    users = q.order_by(User.username).all()
    horas_sorted = sorted({u.weekly_hours for u in users if u.weekly_hours is not None})

    # Obtener centros y categorías dinámicos
    centros = get_centros_dinamicos()
    categorias = get_categorias_disponibles()

    today = date.today().strftime('%Y-%m-%d')
    return render_template("export_excel.html", users=users, today=today, centro_admin=centro_admin, horas_sorted=horas_sorted, centros=centros, categorias=categorias)

# ========== EXCEL DIARIO ==========

@export_bp.route("/excel_daily")
@admin_required
def export_excel_daily():
    fecha_str = request.args.get('fecha', date.today().strftime('%Y-%m-%d'))
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Formato de fecha inválido.", "danger")
        return redirect(url_for("export.export_excel"))

    # Eager loading: cargar User en la misma query para evitar N+1
    records = (
        TimeRecord.query
        .filter(TimeRecord.date == fecha)
        .options(joinedload(TimeRecord.user))
        .order_by(TimeRecord.user_id)
        .all()
    )
    if not records:
        flash("No hay registros para ese día.", "warning")
        return redirect(url_for("export.export_excel"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros diarios"
    header = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Entrada", "Salida", "Horas Trabajadas", "Notas"]
    for col_num, header_text in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for record in records:
        # Usar record.user (ya cargado con eager loading)
        user = record.user
        hours_worked = ""
        if record.check_in and record.check_out:
            time_diff = record.check_out - record.check_in
            hours = time_diff.total_seconds() / 3600
            hours_worked = f"{hours:.2f}"

        ws.cell(row=row_num, column=1).value = user.username if user else f"ID: {record.user_id}"
        ws.cell(row=row_num, column=2).value = user.full_name if user else "-"
        ws.cell(row=row_num, column=3).value = get_user_category_label(user)
        ws.cell(row=row_num, column=4).value = user.center.name if user and user.center else "-"
        ws.cell(row=row_num, column=5).value = record.date.strftime("%d/%m/%Y")
        ws.cell(row=row_num, column=6).value = record.check_in.strftime("%H:%M:%S") if record.check_in else "-"
        ws.cell(row=row_num, column=7).value = record.check_out.strftime("%H:%M:%S") if record.check_out else "-"
        ws.cell(row=row_num, column=8).value = hours_worked
        ws.cell(row=row_num, column=9).value = record.notes
        row_num += 1

    for col_num, _ in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 17

    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(temp_path)

    filename = f"{fecha.strftime('%d_%m_%y')}_TimeT.xlsx"
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========== PDF DIARIO ==========

from fpdf import FPDF

@export_bp.route("/pdf_daily")
@admin_required
def export_pdf_daily():
    fecha_str = request.args.get('fecha', date.today().strftime('%Y-%m-%d'))
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Formato de fecha inválido.", "danger")
        return redirect(url_for("export.export_excel"))

    # Eager loading: cargar User en la misma query para evitar N+1
    records = (
        TimeRecord.query
        .filter(TimeRecord.date == fecha)
        .options(joinedload(TimeRecord.user))
        .order_by(TimeRecord.user_id)
        .all()
    )
    if not records:
        flash("No hay registros para ese día.", "warning")
        return redirect(url_for("export.export_excel"))

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Registros de fichaje del {fecha.strftime('%d/%m/%Y')}", ln=1, align="C")

    pdf.set_font("Arial", "B", 10)
    header = ["Usuario", "Nombre completo", "Categoría", "Centro", "Entrada", "Salida", "Horas", "Notas"]
    col_widths = [30, 40, 25, 30, 22, 22, 30, 55]

    for i, col_name in enumerate(header):
        pdf.cell(col_widths[i], 8, col_name, border=1, align="C")
    pdf.ln()

    pdf.set_font("Arial", "", 9)
    for record in records:
        # Usar record.user (ya cargado con eager loading)
        user = record.user
        hours_worked = ""
        if record.check_in and record.check_out:
            time_diff = record.check_out - record.check_in
            hours = time_diff.total_seconds() / 3600
            hours_worked = f"{hours:.2f}"

        row = [
            user.username if user else f"ID: {record.user_id}",
            user.full_name if user else "-",
            get_user_category_label(user),
            user.center.name if user and user.center else "-",
            record.check_in.strftime("%H:%M:%S") if record.check_in else "-",
            record.check_out.strftime("%H:%M:%S") if record.check_out else "-",
            hours_worked,
            record.notes or ""
        ]
        for i, item in enumerate(row):
            pdf.cell(col_widths[i], 8, str(item), border=1, align="C")
        pdf.ln()

    fd, temp_path = tempfile.mkstemp(suffix='.pdf')
    os.close(fd)
    pdf.output(temp_path)

    filename = f"{fecha.strftime('%d_%m_%y')}_TimeT.pdf"
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )
